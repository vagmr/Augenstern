import os
import csv
import openpyxl
from tkinter import *
from tkinter import filedialog

root = Tk()
root.withdraw()

def choose_file():
    file_path = filedialog.askopenfilename(title='选择一个成绩文件', filetypes=[('text files', '*.txt')])
    if file_path:
        convert_grades(file_path)

def convert_grades(file_path):
    new_file_path = filedialog.asksaveasfilename(title='保存 Excel 文件', defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成绩汇总'

    with open(file_path, 'r') as f:
        reader = csv.reader(f, delimiter=',')
        for i, row in enumerate(reader):
            if i >= 25:  # 只处理前25行
                break
            for j, grade in enumerate(row):
                if not grade.isdigit():
                    continue
                elif int(grade) > 100 or int(grade) < 0:
                    new_grade = '不合格'
                elif int(grade) >= 80:
                    new_grade = '优秀'
                elif int(grade) >= 60:
                    new_grade = '及格'
                else:
                    new_grade = '不及格'
                ws.cell(row=j+1, column=(i-1)%25*2+1, value=grade)
                ws.cell(row=j+1, column=(i-1)%25*2+2, value=new_grade)

    for i in range(1, 51, 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
            
    wb.save(new_file_path)
    print('转换完成，输出文件路径：{}'.format(new_file_path))

choose_file()
